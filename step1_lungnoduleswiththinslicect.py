''' step 1のメイン実行ファイル
全データの入っているExcelファイルを読み込み、thin slice CTが撮影されている
肺結節のみに絞り込み、結果ファイルに書き出す。
'''

import json
import os.path

import openpyxl

__author__ = 'Takeyuki Watadani<twatadani@g.ecc.u-tokyo.ac.jp>'
__version__ = '0.2'
__date__ = '2021/10/25'
__status__ = 'dev'

print('step 1を開始します。')

# configの読み込み
with open('config.json', mode='rt') as fp:
    cfdict = json.load(fp)

# 生データの読み込み
datadir = os.path.expanduser(os.path.expandvars(cfdict['datadir']))
rawfile = os.path.join(datadir, cfdict['rawdata'])

workbook = openpyxl.load_workbook(rawfile)
sheet = workbook['シート1']

# 新しいデータファイルの作成
newwb = openpyxl.Workbook()
ws = newwb.active
ws.title = 'Sheet1'


# thin slice CTが撮影されている結節の行のみを抽出する
rows = sheet.iter_rows()
dstrow = 1
for i, row in enumerate(rows, 1):
    # 最初の行は必ずコピー
    if i == 1:
        copy_flag = True
    else:
        nodule_index = 2
        thinslice_index = 7
        if row[nodule_index].value == 1 and row[thinslice_index].value == 1:
            copy_flag = True
        else:
            copy_flag = False

    if copy_flag is True:
        for col in range(17):
            _ = ws.cell(column=col+1, row=dstrow, value=row[col].value)
        dstrow += 1
    
# 新しいワークブックを書き出す
dstfile = cfdict['step1resultfile']
dstfilename = os.path.join(datadir, dstfile)
newwb.save(filename = dstfilename)

print('step 1が終了しました。')