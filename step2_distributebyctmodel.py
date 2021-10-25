''' step2_distributebyctmodel.py - step 1で絞り込んだ症例群をCTモデルと良悪性ごとに仕分けする '''

import json
import os.path

import openpyxl

__author__ = 'Takeyuki Watadani<twatadani@g.ecc.u-tokyo.ac.jp>'
__version__ = '0.2'
__date__ = '2021/10/25'
__status__ = 'dev'

print('step 2を開始します。')

# configの読み込み
with open('config.json', mode='rt') as fp:
    cfdict = json.load(fp)

# step 1結果の読み込み

# 生データの読み込み
datadir = os.path.expanduser(os.path.expandvars(cfdict['datadir']))
srcfile = os.path.join(datadir, cfdict['step1resultfile'])

workbook = openpyxl.load_workbook(srcfile)
sheet = workbook['Sheet1']

# CTモデルのリスト

CTmodels = {
    'Aquilion': 'aquilion',
    'Aquilion PRIME': 'aquilion_prime',
    'Aquilion ONE': 'aquilion_one',
    'Aquilion Precision': 'aquilion_precision',
    'Revolution CT': 'revolution_ct',
    'Discovery CT750 HD': 'discovery_ct750_hd',
    'Discovery CT751 HD': 'discovery_ct751_hd'
    }

# CTモデルの判定関数

def define_model(srcstr: str):
    if '他院' in srcstr:
        return None
    elif 'SOMATOM' in srcstr:
        return None
    elif 'ONE' in srcstr:
        return 'Aquilion ONE'
    elif 'PRIME' in srcstr:
        return 'Aquilion PRIME'
    elif 'プレシジョン' in srcstr:
        return 'Aquilion Precision'
    elif 'Precision' in srcstr:
        return 'Aquilion Precision'
    elif 'Aquilion' in srcstr:
        return 'Aquilion'
    elif 'Revolution' in srcstr:
        return 'Revolution CT'
    elif 'CT750' in srcstr:
        return 'Discovery CT750 HD'
    elif 'CT751' in srcstr:
        return 'Discovery CT751 HD'
    else:
        print('CTの機種文字列の対応が不明です:', srcstr)
        raise AssertionError

# 新しいワークブックとシートを準備
wbdict = {}
for model in CTmodels:
    wbdict[model] = openpyxl.Workbook()
    ws_benign = wbdict[model].active
    ws_benign.title = 'Benign'
    wbdict[model].create_sheet(title='Malignant')

# ソースファイルを走査してコピー

# コピー用の行番号
dstrows = {}
for model in CTmodels:
    dstrows[model] = {}
    dstrows[model]['Benign'] = 1
    dstrows[model]['Malignant'] = 1

for i, row in enumerate(sheet.iter_rows(), 1):
    if i == 1: # 先頭行
        for model in CTmodels:
            wb = wbdict[model]
            ws_benign = wb['Benign']
            ws_malignant = wb['Malignant']
            for col in range(17):
                _ = ws_benign.cell(column=col+1, row=dstrows[model]['Benign'], value=row[col].value)
                _ = ws_malignant.cell(column=col+1, row=dstrows[model]['Malignant'], value=row[col].value)
            dstrows[model]['Benign'] = dstrows[model]['Malignant'] = dstrows[model]['Benign'] + 1
    else:
        modelstr = row[9].value
        refined_model = define_model(modelstr)
        if refined_model is not None:
            # print(row[11])
            malignancy = 'Malignant' if row[11].value == 1 else 'Benign'
            dstws = wbdict[refined_model][malignancy]
            # 書き込み
            for col in range(17):
                _ = dstws.cell(column=col+1, row=dstrows[refined_model][malignancy], value=row[col].value)
            dstrows[refined_model][malignancy] += 1

# 新しいワークブックを保存
for model in CTmodels:
    wbfilename = os.path.join(datadir, cfdict['step2resultfile'] + CTmodels[model] + '.xlsx')
    wb = wbdict[model]
    wb.save(wbfilename)

print('step 2を終了します。')